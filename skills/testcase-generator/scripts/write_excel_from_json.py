#!/usr/bin/env python3
"""
从JSON文件读取测试用例并写入Excel
解决命令行长度超限问题
"""

import argparse
import json
import os
import sys
import zipfile
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("❌ 缺少 openpyxl 库，请运行: pip install openpyxl")
    sys.exit(1)


def get_config(templates_dir):
    """加载模板配置"""
    config_path = templates_dir / "template-config.json"
    if config_path.exists():
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None


def read_xlsx_template_columns(template_path):
    """从Excel模板读取列定义"""
    columns = []
    with zipfile.ZipFile(template_path, 'r') as z:
        shared_strings = []
        try:
            with z.open('xl/sharedStrings.xml') as f:
                tree = ET.parse(f) if 'ET' in dir() else None
        except:
            pass
        
        # 直接用openpyxl读取更简单
        pass
    
    # 使用openpyxl直接读取
    from openpyxl import load_workbook
    wb = load_workbook(template_path)
    ws = wb.active
    
    for cell in ws[1]:
        if cell.value:
            columns.append({
                "index": cell.column - 1,
                "name": cell.value
            })
    
    return columns


def merge_columns(xlsx_columns, config_columns):
    """合并Excel列头和配置列定义"""
    name_to_key = {col['name']: col['key'] for col in config_columns}
    name_to_required = {col['name']: col.get('required', False) for col in config_columns}
    
    result = []
    for col in xlsx_columns:
        key = name_to_key.get(col['name'], col['name'].lower().replace(' ', '_'))
        required = name_to_required.get(col['name'], False)
        result.append({
            "index": col['index'],
            "name": col['name'],
            "key": key,
            "required": required
        })
    return result


def format_steps(value):
    """将分隔的步骤转换为多行格式"""
    if not isinstance(value, str):
        return value
    # 处理HTML换行符 <br> → \n
    value = value.replace('<br>', '\n').replace('<br/>', '\n').replace('<BR>', '\n')

    # 如果已经是多行，直接清理
    if '\n' in value:
        lines = [line.strip() for line in value.split('\n') if line.strip()]
        return '\n'.join(lines)

    # 如果包含分号（中文或英文），将分号替换为换行
    if '；' in value or ';' in value:
        value = value.replace('；', '\n').replace(';', '\n')
        lines = [line.strip() for line in value.split('\n') if line.strip()]
        return '\n'.join(lines)

    # 如果是数字序号格式（如 "1. xxx 2. xxx 3. xxx"），按序号拆分
    import re
    # 匹配 "数字. " 格式
    pattern = r'(\d+[.、])\s*'
    # 在每个 "数字. " 前面插入换行符
    value = re.sub(pattern, r'\n\1', value).strip()
    if value.startswith('\n'):
        value = value[1:]

    return value


def write_excel(output_path, columns, testcases, title="测试用例"):
    """写入Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # 写入表头
    for col in columns:
        cell = ws.cell(row=1, column=col['index'] + 1, value=col['name'])
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    for row_idx, tc in enumerate(testcases, 2):
        for col in columns:
            key = col['key']
            value = tc.get(key, "")
            # 处理列表/字典类型
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            # 格式化步骤（分号转换行）
            value = format_steps(value)
            cell = ws.cell(row=row_idx, column=col['index'] + 1, value=str(value) if value else "")
            cell.border = thin_border
            cell.alignment = wrap_alignment
    
    # 设置列宽
    for col in columns:
        col_letter = chr(64 + col['index'] + 1)
        ws.column_dimensions[col_letter].width = 20
    
    ws.freeze_panes = 'A2'
    wb.save(output_path)
    return len(testcases)


def main():
    parser = argparse.ArgumentParser(description='从JSON文件读取测试用例写入Excel')
    parser.add_argument('--data', '-d', required=True, help='JSON文件路径（包含testcases数组）')
    parser.add_argument('--template', '-t', default='测试用例模板.xlsx', help='Excel模板名称')
    parser.add_argument('--output', '-o', required=True, help='输出Excel路径')
    parser.add_argument('--title', default='测试用例', help='工作表标题')
    parser.add_argument('--template-key', default='default', help='template-config.json中的模板key')
    
    args = parser.parse_args()
    
    # 检查JSON文件
    if not os.path.exists(args.data):
        print(f"❌ JSON文件不存在: {args.data}")
        sys.exit(1)
    
    # 处理路径
    script_dir = Path(os.path.dirname(os.path.abspath(__file__)))
    templates_dir = script_dir.parent / 'assets'
    
    # 1. 加载JSON数据
    print(f"📄 读取数据文件: {args.data}")
    with open(args.data, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 支持直接是数组或 {testcases: [...]} 格式
    testcases = data if isinstance(data, list) else data.get('testcases', [])
    print(f"   测试用例数: {len(testcases)}")
    
    # 2. 加载配置
    config = get_config(templates_dir)
    if not config:
        print("❌ 未找到 template-config.json")
        sys.exit(1)
    
    config_columns = config.get('templates', {}).get(args.template_key, {}).get('columns', [])
    
    # 3. 读取Excel模板
    template_path = templates_dir / args.template
    if not template_path.exists():
        print(f"❌ 模板文件不存在: {template_path}")
        sys.exit(1)
    
    xlsx_columns = read_xlsx_template_columns(str(template_path))
    columns = merge_columns(xlsx_columns, config_columns)
    print(f"   模板列数: {len(columns)}")
    
    # 4. 写入Excel
    output_path = os.path.abspath(args.output)
    count = write_excel(output_path, columns, testcases, args.title)
    
    print(f"\n✅ Excel已生成: {output_path}")
    print(f"   测试用例数: {count}")


if __name__ == '__main__':
    main()
