#!/usr/bin/env python3
"""
读取Excel测试用例文件，提取结构化数据
"""

import json
import sys
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def read_excel_cases(file_path: str) -> dict:
    """
    读取Excel测试用例文件

    Args:
        file_path: Excel文件路径

    Returns:
        包含modules和testcases的字典
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # 获取表头行
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # 标准化表头映射
    header_map = {}
    for idx, header in enumerate(headers):
        if header:
            header_lower = str(header).strip().lower()
            if '编号' in header_lower or 'id' in header_lower:
                header_map['case_id'] = idx
            elif '模块' in header_lower or '所属' in header_lower:
                header_map['module'] = idx
            elif '标题' in header_lower or '名称' in header_lower:
                header_map['title'] = idx
            elif '类型' in header_lower:
                header_map['type'] = idx
            elif '优先级' in header_lower or '优先' in header_lower:
                header_map['priority'] = idx
            elif '步骤' in header_lower:
                header_map['steps'] = idx
            elif '预期' in header_lower:
                header_map['expected'] = idx

    # 提取数据行
    testcases = []
    modules = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue

        case = {
            'case_id': row[header_map.get('case_id', 0)] if header_map.get('case_id') is not None else f'TC_{row_idx:03d}',
            'module': str(row[header_map.get('module', 1)]).strip() if header_map.get('module') is not None and row[header_map.get('module', 1)] else '未分类',
            'title': str(row[header_map.get('title', 2)]).strip() if header_map.get('title') is not None and row[header_map.get('title', 2)] else '',
            'type': str(row[header_map.get('type', 3)]).strip() if header_map.get('type') is not None and row[header_map.get('type', 3)] else '功能测试',
            'priority': str(row[header_map.get('priority', 4)]).strip() if header_map.get('priority') is not None and row[header_map.get('priority', 4)] else 'P2',
            'steps': str(row[header_map.get('steps', 6)]).strip() if header_map.get('steps') is not None and row[header_map.get('steps', 6)] else '',
            'expected': str(row[header_map.get('expected', 7)]).strip() if header_map.get('expected') is not None and row[header_map.get('expected', 7)] else ''
        }

        if case['module']:
            modules.add(case['module'])

        testcases.append(case)

    return {
        'modules': sorted(list(modules)),
        'testcases': testcases
    }


def main():
    parser = argparse.ArgumentParser(description='读取Excel测试用例文件')
    parser.add_argument('file_path', help='Excel文件路径')
    parser.add_argument('-o', '--output', help='输出JSON文件路径')

    args = parser.parse_args()

    if not Path(args.file_path).exists():
        print(f"Error: File not found: {args.file_path}")
        sys.exit(1)

    result = read_excel_cases(args.file_path)

    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"Output written to: {args.output}")
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
